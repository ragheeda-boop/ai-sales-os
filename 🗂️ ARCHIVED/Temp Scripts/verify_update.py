import openpyxl

excel_path = r"C:\Users\PC\Documents\AI Sales OS\📊 DATA\APOLLO_NOTION_FIELD_GAP_ANALYSIS.xlsx"
output_path = r"C:\Users\PC\Documents\AI Sales OS\VERIFICATION_RESULT.txt"

wb = openpyxl.load_workbook(excel_path)

with open(output_path, 'w') as f:
    f.write("=== VERIFICATION OF UPDATES ===\n\n")
    
    # Check Critical Bugs
    f.write("Critical Bugs Sheet:\n")
    bugs_sheet = wb["Critical Bugs"]
    for row in bugs_sheet.iter_rows():
        cell_val = row[0].value
        if cell_val and isinstance(cell_val, str) and "BUG-" in cell_val:
            f.write(f"  {cell_val}: Status = {row[1].value}\n")
    
    # Check Contact Fields
    f.write("\nContact Fields Gap (updated rows):\n")
    contacts_sheet = wb["Contact Fields Gap"]
    check_rows = [24, 36, 37, 38, 39, 40, 41, 42]
    for row_num in check_rows:
        row = contacts_sheet[row_num]
        f.write(f"  Row {row_num}: {row[0].value}\n")
        f.write(f"    Status: {row[3].value}\n")
        f.write(f"    Note: {row[4].value}\n")
    
    # Check Company Fields
    f.write("\nCompany Fields Gap (Account Stage):\n")
    companies_sheet = wb["Company Fields Gap"]
    for row in companies_sheet.iter_rows():
        if row[0].value == "Account Stage":
            f.write(f"  Status: {row[3].value}\n")
            f.write(f"  Note: {row[4].value}\n")
            break
    
    # Check Summary
    f.write("\nSummary Sheet (last 2 rows):\n")
    summary_sheet = wb["Summary"]
    for i in range(summary_sheet.max_row - 1, summary_sheet.max_row + 1):
        cell = summary_sheet[i][0].value
        if cell:
            f.write(f"  Row {i}: {cell}\n")

print("Verification complete - check VERIFICATION_RESULT.txt")
