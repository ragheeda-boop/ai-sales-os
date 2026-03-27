import openpyxl
excel_path = r"C:\Users\PC\Documents\AI Sales OS\📊 DATA\APOLLO_NOTION_FIELD_GAP_ANALYSIS.xlsx"
wb = openpyxl.load_workbook(excel_path)

# Critical Bugs
bugs_sheet = wb["Critical Bugs"]
for row in bugs_sheet.iter_rows():
    cell_val = row[0].value
    if cell_val and isinstance(cell_val, str):
        for bug_num in [5, 6, 7, 8, 9]:
            if f"BUG-{bug_num}" in cell_val:
                row[1].value = "FIXED"
                break

# Contact Fields
contacts_sheet = wb["Contact Fields Gap"]
rows_to_update = {
    24: ("OK", "Fixed in v2.1 — synced as checkbox via safe boolean writing"),
    40: ("OK", "Fixed in v2.1 — Notion property added, synced via safe boolean"),
    41: ("OK", "Fixed in v2.1 — now written as date field"),
    42: ("OK", "Fixed in v2.1 — synced as select field"),
    36: ("OK", "Fixed in v2.1 — safe boolean writing prevents false overwrite"),
    37: ("OK", "Fixed in v2.1 — safe boolean writing"),
    38: ("OK", "Fixed in v2.1 — safe boolean writing"),
    39: ("OK", "Fixed in v2.1 — safe boolean writing")
}

for row_num, (status, note) in rows_to_update.items():
    row = contacts_sheet[row_num]
    row[3].value = status
    row[4].value = note

# Company Fields
companies_sheet = wb["Company Fields Gap"]
for row in companies_sheet.iter_rows():
    if row[0].value == "Account Stage":
        row[3].value = "OK"
        row[4].value = "Fixed in v2.1 — synced as select field"
        break

# Summary
summary_sheet = wb["Summary"]
summary_sheet.append([])
summary_sheet.append(["Updated 27 March 2026 — reflects daily_sync.py v2.1 fixes"])

wb.save(excel_path)
print("Updated successfully")
