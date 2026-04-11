#!/usr/bin/env python
"""Update APOLLO_NOTION_FIELD_GAP_ANALYSIS.xlsx with v2.1 bug fixes."""

import openpyxl
import sys

excel_path = r"C:\Users\PC\Documents\AI Sales OS\📊 DATA\APOLLO_NOTION_FIELD_GAP_ANALYSIS.xlsx"
log_path = r"C:\Users\PC\Documents\AI Sales OS\update_gap_analysis.log"

log_file = open(log_path, 'w')

def log(msg):
    print(msg)
    log_file.write(msg + '\n')
    log_file.flush()

try:
    log("Loading workbook...")
    wb = openpyxl.load_workbook(excel_path)
    log(f"Sheets: {wb.sheetnames}")
    
    # Critical Bugs
    log("\nUpdating Critical Bugs...")
    bugs_sheet = wb["Critical Bugs"]
    for bug_num in [5, 6, 7, 8, 9]:
        for row in bugs_sheet.iter_rows():
            if row[0].value and f"BUG-{bug_num}" in str(row[0].value):
                row[1].value = "FIXED"
                log(f"  BUG-{bug_num}: FIXED")
                break
    
    # Contact Fields
    log("\nUpdating Contact Fields...")
    contacts_sheet = wb["Contact Fields Gap"]
    updates = [
        (24, "OK", "Fixed in v2.1 — synced as checkbox via safe boolean writing"),
        (40, "OK", "Fixed in v2.1 — Notion property added, synced via safe boolean"),
        (41, "OK", "Fixed in v2.1 — now written as date field"),
        (42, "OK", "Fixed in v2.1 — synced as select field"),
        (36, "OK", "Fixed in v2.1 — safe boolean writing prevents false overwrite"),
        (37, "OK", "Fixed in v2.1 — safe boolean writing"),
        (38, "OK", "Fixed in v2.1 — safe boolean writing"),
        (39, "OK", "Fixed in v2.1 — safe boolean writing")
    ]
    
    for row_num, status, note in updates:
        row = contacts_sheet[row_num]
        row[3].value = status
        row[4].value = note
        log(f"  Row {row_num}: {status}")
    
    # Company Fields
    log("\nUpdating Company Fields...")
    companies_sheet = wb["Company Fields Gap"]
    for row in companies_sheet.iter_rows():
        if row[0].value == "Account Stage":
            row[3].value = "OK"
            row[4].value = "Fixed in v2.1 — synced as select field"
            log("  Account Stage: OK")
            break
    
    # Summary
    log("\nUpdating Summary...")
    summary_sheet = wb["Summary"]
    summary_sheet.append([])
    summary_sheet.append(["Updated 27 March 2026 — reflects daily_sync.py v2.1 fixes"])
    log("  Timestamp added")
    
    # Save
    log("\nSaving...")
    wb.save(excel_path)
    log("SUCCESS")
    
except Exception as e:
    log(f"ERROR: {e}")
    import traceback
    log(traceback.format_exc())
    
finally:
    log_file.close()
