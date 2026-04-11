from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define colors
ORANGE_THEME = "E8771A"
HEADER_FILL = PatternFill(start_color=ORANGE_THEME, end_color=ORANGE_THEME, fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)

# Status colors
STATUS_COLORS = {
    "OK": "C6EFCE",  # Light green
    "PARTIAL": "FFEB9C",  # Light yellow
    "MISSING": "FFC7CE",  # Light red
    "SKIPPED": "D9D9D9",  # Light gray
    "FIXED": "C6EFCE",  # Light green
    "OPEN": "FFE699",  # Light orange
}

def apply_formatting(ws, headers, data_rows=None):
    """Apply header and body formatting"""
    # Header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data rows
    if data_rows:
        for row_num, row_data in enumerate(data_rows, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = BODY_FONT
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Alternate row shading
                if row_num % 2 == 0:
                    cell.fill = ALT_ROW_FILL
    
    # Auto-width columns
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 18

def color_status_cells(ws, status_col_index):
    """Color code Status column"""
    for row_num in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=status_col_index)
        status = cell.value
        if status in STATUS_COLORS:
            cell.fill = PatternFill(start_color=STATUS_COLORS[status], 
                                   end_color=STATUS_COLORS[status], 
                                   fill_type="solid")

# ============ SHEET 1: Contact Fields Gap ============
ws1 = wb.create_sheet("Contact Fields Gap")
contact_headers = ["#", "Apollo Export Field", "In Sync Code?", "API Field Name", 
                   "In Notion?", "Notion Property Name", "Notion Type", "Status", "Impact", "Notes"]
contact_data = [
    [1, "First Name", "YES", "first_name", "YES", "First Name", "rich_text", "OK", "—", ""],
    [2, "Last Name", "YES", "last_name", "YES", "Last Name", "rich_text", "OK", "—", ""],
    [3, "(Full Name)", "YES", "name", "YES", "Full Name", "title", "OK", "—", ""],
    [4, "Email", "YES", "email", "YES", "Email", "email", "OK", "—", ""],
    [5, "Email Status", "YES", "email_status", "YES", "Email Status", "select", "OK", "—", ""],
    [6, "Secondary Email", "YES", "secondary_email", "YES", "Secondary Email", "email", "OK", "—", ""],
    [7, "Work Direct Phone", "YES", "direct_phone/phone", "YES", "Work Direct Phone", "phone_number", "OK", "—", ""],
    [8, "Home Phone", "YES", "home_phone", "YES", "Home Phone", "phone_number", "OK", "—", ""],
    [9, "Mobile Phone", "YES", "mobile_phone", "YES", "Mobile Phone", "phone_number", "OK", "—", ""],
    [10, "Corporate Phone", "YES", "corporate_phone", "YES", "Corporate Phone", "phone_number", "OK", "—", ""],
    [11, "Other Phone", "YES", "other_phone", "YES", "Other Phone", "phone_number", "OK", "—", ""],
    [12, "Do Not Call", "YES", "do_not_call", "YES", "Do Not Call", "checkbox", "OK", "—", "Fixed v2.1 — safe boolean writing"],
    [13, "Title", "YES", "title", "YES", "Title", "rich_text", "OK", "—", ""],
    [14, "Seniority", "YES", "seniority", "YES", "Seniority", "select", "OK", "—", ""],
    [15, "Departments", "YES", "departments", "YES", "Departments", "rich_text", "OK", "—", ""],
    [16, "City", "YES", "city", "YES", "City", "rich_text", "OK", "—", ""],
    [17, "State", "YES", "state", "YES", "State", "rich_text", "OK", "—", ""],
    [18, "Country", "YES", "country", "YES", "Country", "rich_text", "OK", "—", ""],
    [19, "Person Linkedin Url", "YES", "linkedin_url", "YES", "Person Linkedin Url", "url", "OK", "—", ""],
    [20, "Apollo Contact Id", "YES", "id", "YES", "Apollo Contact Id", "rich_text", "OK", "—", ""],
    [21, "Apollo Account Id", "YES", "account_id", "YES", "Apollo Account Id", "rich_text", "OK", "—", ""],
    [22, "Email Sent", "YES", "email_sent", "YES", "Email Sent", "checkbox", "OK", "—", "Fixed v2.1 — safe boolean writing"],
    [23, "Email Opened", "YES", "email_open", "YES", "Email Opened", "checkbox", "OK", "—", "Fixed v2.1 — safe boolean writing"],
    [24, "Email Bounced", "YES", "email_bounced", "YES", "Email Bounced", "checkbox", "OK", "—", "Fixed v2.1 — safe boolean writing"],
    [25, "Replied", "YES", "replied", "YES", "Replied", "checkbox", "OK", "—", "Fixed v2.1 — safe boolean writing"],
    [26, "Meeting Booked", "YES", "meeting_booked", "YES", "Meeting Booked", "checkbox", "OK", "—", "Fixed v2.1 — safe boolean writing"],
    [27, "Demoed", "YES", "demoed", "YES", "Demoed", "checkbox", "OK", "—", "Fixed v2.1 — Notion property added"],
    [28, "Last Contacted", "YES", "last_activity_date", "YES", "Last Contacted", "date", "OK", "—", "Fixed v2.1 — written as date"],
    [29, "Stage", "YES", "stage", "YES", "Stage", "select", "OK", "—", "Fixed v2.1 — synced as select"],
    [30, "Outreach Status", "YES", "outreach_status", "YES", "Outreach Status", "select", "OK", "—", "Added in v2.1"],
    [31, "Contact Owner", "YES", "owner_id", "NO", "—", "—", "PARTIAL", "MED", "Pulled but not sent to Notion"],
    [32, "Company Name", "YES", "organization_name", "YES", "Company Name for Emails", "rich_text", "PARTIAL", "MED", "In dict but not written as text field"],
    [33, "Primary Email Source", "NO", "—", "NO", "—", "—", "MISSING", "LOW", "Not in API pull"],
    [34, "Email Confidence", "NO", "—", "NO", "—", "—", "SKIPPED", "—", "Intentionally skipped"],
    [35, "Lists", "NO", "—", "NO", "—", "—", "MISSING", "MED", "Apollo list membership"],
]
apply_formatting(ws1, contact_headers, contact_data)
color_status_cells(ws1, 8)  # Status column (8)

# ============ SHEET 2: Company Fields Gap ============
ws2 = wb.create_sheet("Company Fields Gap")
company_headers = ["#", "Apollo Export Field", "In Sync Code?", "API Field Name", 
                   "In Notion?", "Notion Property Name", "Notion Type", "Status", "Impact", "Notes"]
company_data = [
    [1, "Company Name", "YES", "name", "YES", "Company Name", "title", "OK", "—", ""],
    [2, "Apollo Account Id", "YES", "id", "YES", "Apollo Account Id", "rich_text", "OK", "—", ""],
    [3, "Domain", "YES", "domain", "YES", "Domain", "rich_text", "OK", "—", "Fixed v2.1 — now written"],
    [4, "Website", "YES", "website_url", "YES", "Website", "url", "OK", "—", "Fixed v2.1 — now written"],
    [5, "Industry", "YES", "industry", "YES", "Industry", "rich_text", "OK", "—", "Fixed v2.1 — now written"],
    [6, "Employees", "YES", "employee_count", "YES", "Employees", "number", "OK", "—", "Fixed v2.1 — now written"],
    [7, "Employee Size", "YES", "employee_size_category", "YES", "Employee Size", "rich_text", "OK", "—", "Fixed v2.1"],
    [8, "Annual Revenue", "YES", "annual_revenue", "YES", "Annual Revenue", "number", "OK", "—", "Fixed v2.1 — now written"],
    [9, "Revenue Range", "YES", "revenue_range", "YES", "Revenue Range", "rich_text", "OK", "—", "Fixed v2.1"],
    [10, "Total Funding", "YES", "total_funding", "YES", "Total Funding", "number", "OK", "—", "Fixed v2.1"],
    [11, "Latest Funding Amount", "YES", "latest_funding_amount", "YES", "Latest Funding Amount", "number", "OK", "—", "Fixed v2.1"],
    [12, "Last Raised At", "YES", "last_funding_date", "YES", "Last Raised At", "date", "OK", "—", "Fixed v2.1"],
    [13, "Account Stage", "YES", "account_stage", "YES", "Account Stage", "select", "OK", "—", "Fixed v2.1"],
    [14, "Company City", "YES", "hq_city", "YES", "Company City", "rich_text", "OK", "—", "Fixed v2.1"],
    [15, "Company State", "YES", "hq_state", "YES", "Company State", "rich_text", "OK", "—", "Fixed v2.1"],
    [16, "Company Country", "YES", "hq_country", "YES", "Company Country", "rich_text", "OK", "—", "Fixed v2.1"],
    [17, "Company Address", "YES", "hq_address", "YES", "Company Address", "rich_text", "OK", "—", "Fixed v2.1"],
    [18, "Company Phone", "YES", "phone", "YES", "Company Phone", "phone_number", "OK", "—", "Fixed v2.1"],
    [19, "Company Linkedin Url", "YES", "linkedin_url", "YES", "Company Linkedin Url", "url", "OK", "—", "Fixed v2.1"],
    [20, "Facebook Url", "YES", "facebook_url", "YES", "Facebook Url", "url", "OK", "—", "Fixed v2.1"],
    [21, "Twitter Url", "YES", "twitter_url", "YES", "Twitter Url", "url", "OK", "—", "Fixed v2.1"],
    [22, "Keywords", "YES", "keywords", "YES", "Keywords", "rich_text", "OK", "—", "Fixed v2.1"],
    [23, "Technologies", "YES", "technologies", "YES", "Technologies", "rich_text", "OK", "—", "Fixed v2.1"],
    [24, "Short Description", "NO", "—", "YES", "Short Description", "rich_text", "MISSING", "LOW", "Not in API pull"],
    [25, "Account Owner", "YES", "account_owner_id", "YES", "Account Owner", "rich_text", "PARTIAL", "MED", "In dict but not sent"],
]
apply_formatting(ws2, company_headers, company_data)
color_status_cells(ws2, 8)  # Status column (8)

# ============ SHEET 3: Bugs Fixed ============
ws3 = wb.create_sheet("Bugs Fixed")
bugs_headers = ["#", "Bug ID", "Description", "Severity", "Status", "Fix Details"]
bugs_data = [
    [1, "BUG-1", "create_company used 'Name' instead of 'Company Name'", "CRITICAL", "FIXED", "Fixed in daily_sync.py v2.1"],
    [2, "BUG-2", "create/update_company only sent Apollo ID", "CRITICAL", "FIXED", "All company fields now written in v2.1"],
    [3, "BUG-5", "Engagement booleans defaulted to False", "HIGH", "FIXED", "Safe boolean writing in v2.1"],
    [4, "BUG-6", "Demoed field not in Notion", "HIGH", "FIXED", "Property added, synced via safe boolean"],
    [5, "BUG-7", "Stage field completely missing", "HIGH", "FIXED", "Synced as select field in v2.1"],
    [6, "BUG-8", "Do Not Call flag missing", "HIGH", "FIXED", "Synced as checkbox via safe boolean"],
    [7, "BUG-9", "Last Contacted not written to Notion", "MEDIUM", "FIXED", "Written as date field from last_activity_date"],
    [8, "BUG-10", "Contact Owner pulled but not stored", "MEDIUM", "OPEN", "Still not mapped to Notion property"],
]
apply_formatting(ws3, bugs_headers, bugs_data)
color_status_cells(ws3, 5)  # Status column (5)

# ============ SHEET 4: Summary ============
ws4 = wb.create_sheet("Summary")
summary_headers = ["Metric", "Count"]
summary_data = [
    ["Contact Fields OK", 29],
    ["Contact Fields PARTIAL", 2],
    ["Contact Fields MISSING", 3],
    ["Contact Fields SKIPPED", 1],
    ["Company Fields OK", 23],
    ["Company Fields PARTIAL", 1],
    ["Company Fields MISSING", 1],
    ["Bugs Fixed", "7 of 8"],
    ["Last Updated", "27 March 2026"],
]
apply_formatting(ws4, summary_headers, summary_data)

# Save workbook
output_path = r"C:\Users\PC\Documents\AI Sales OS\📊 DATA\APOLLO_NOTION_FIELD_GAP_ANALYSIS.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
